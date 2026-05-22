"""数据源扩展。

支持从 Excel (.xlsx) 文件加载数据驱动测试数据。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any


def load_excel_data(file_path: str | Path, sheet: str | None = None) -> list[dict[str, Any]]:
    """从 Excel 文件加载数据。

    Args:
        file_path: Excel 文件路径
        sheet: 工作表名称，默认第一个

    Returns:
        数据行列表，每行为 dict
    """
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:
        raise ImportError("Excel 数据源需要 openpyxl，请运行 uv add openpyxl") from exc

    path = Path(file_path)
    if not path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet:
        if sheet not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet}' 不存在，可用: {wb.sheetnames}")
        ws = wb[sheet]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        entry = {}
        for i, value in enumerate(row):
            if i < len(headers):
                entry[headers[i]] = value
        data.append(entry)

    wb.close()
    return data


def load_sql_data(
    connection_string: str,
    query: str,
) -> list[dict[str, Any]]:
    """从数据库 SQL 查询加载数据（骨架）。

    首期仅提供接口定义，完整实现需后续补齐数据库驱动。
    """
    raise NotImplementedError(
        "SQL 数据源将在后续版本实现。当前可通过 CSV/JSON 数据文件或 hook 函数实现数据库数据注入。"
    )
